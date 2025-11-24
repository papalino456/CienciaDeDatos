"""
Script to organize JSONL data files into a readable Excel (.xlsx) file.

This script reads all JSONL files from the data directory and organizes them
into a well-formatted Excel workbook with separate sheets for each data file
and a summary sheet with statistics.
"""

import json
import os
from pathlib import Path
from collections import Counter
from typing import Dict, List, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm


def read_jsonl(file_path: Path) -> List[Dict[str, Any]]:
    """Read a JSONL file and return a list of dictionaries."""
    data = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:
                    data.append(json.loads(line))
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
    return data


def get_all_jsonl_files(data_dir: Path) -> List[Path]:
    """Find all JSONL files in the data directory recursively."""
    jsonl_files = []
    for root, dirs, files in os.walk(data_dir):
        # Skip tokenizer directory
        if 'tokenizer' in root:
            continue
        for file in files:
            if file.endswith('.jsonl'):
                jsonl_files.append(Path(root) / file)
    return sorted(jsonl_files)


def create_summary_sheet(writer, all_data: Dict[str, List[Dict]]):
    """Create a summary sheet with statistics about all datasets."""
    summary_data = []
    
    for sheet_name, data in all_data.items():
        if not data:
            continue
            
        df = pd.DataFrame(data)
        
        # Basic statistics
        total_records = len(data)
        topics = df['topic'].value_counts().to_dict() if 'topic' in df.columns else {}
        unique_urls = df['url'].nunique() if 'url' in df.columns else 0
        
        # Text statistics
        if 'text' in df.columns:
            text_lengths = df['text'].str.len()
            avg_text_length = text_lengths.mean()
            min_text_length = text_lengths.min()
            max_text_length = text_lengths.max()
        else:
            avg_text_length = min_text_length = max_text_length = 0
        
        summary_data.append({
            'Dataset': sheet_name,
            'Total Records': total_records,
            'Unique URLs': unique_urls,
            'Avg Text Length': round(avg_text_length, 1),
            'Min Text Length': int(min_text_length),
            'Max Text Length': int(max_text_length),
            'Topics': len(topics),
            'Topic Distribution': ', '.join([f"{k}: {v}" for k, v in sorted(topics.items(), key=lambda x: -x[1])[:5]])
        })
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)


def format_summary_sheet(workbook):
    """Format the summary sheet with proper styling."""
    worksheet = workbook['Summary']
    
    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data row formatting
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def format_data_sheet(workbook, sheet_name: str):
    """Format a data sheet with proper styling."""
    worksheet = workbook[sheet_name]
    
    # Header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data row formatting
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    # For text column, limit the length check
                    if column_letter == 'A':  # Assuming text is in column A
                        max_length = max(max_length, min(len(str(cell.value)), 100))
                    else:
                        max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 100 if column_letter == 'A' else 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'


def organize_data_to_xlsx(data_dir: Path, output_path: Path):
    """Main function to organize all JSONL files into an Excel workbook."""
    print(f"Scanning for JSONL files in {data_dir}...")
    jsonl_files = get_all_jsonl_files(data_dir)
    
    if not jsonl_files:
        print("No JSONL files found!")
        return
    
    print(f"Found {len(jsonl_files)} JSONL file(s):")
    for f in jsonl_files:
        print(f"  - {f}")
    
    # Read all data files
    all_data = {}
    for jsonl_file in tqdm(jsonl_files, desc="Reading JSONL files"):
        # Create a clean sheet name from the file path
        # e.g., "data/splits/test.jsonl" -> "splits_test"
        relative_path = jsonl_file.relative_to(data_dir)
        sheet_name = str(relative_path).replace(os.sep, '_').replace('.jsonl', '')
        # Excel sheet names are limited to 31 characters
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:28] + "..."
        
        data = read_jsonl(jsonl_file)
        if data:
            all_data[sheet_name] = data
            print(f"  Loaded {len(data)} records from {jsonl_file.name}")
    
    if not all_data:
        print("No data found in any JSONL files!")
        return
    
    # Create Excel file
    print(f"\nCreating Excel file: {output_path}")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write each dataset to a separate sheet
        for sheet_name, data in tqdm(all_data.items(), desc="Writing sheets"):
            df = pd.DataFrame(data)
            # Ensure columns are in a logical order: text, topic, url
            if 'text' in df.columns and 'topic' in df.columns and 'url' in df.columns:
                df = df[['text', 'topic', 'url']]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Create summary sheet
        create_summary_sheet(writer, all_data)
    
    # Format all sheets (after the file is fully written)
    print("Formatting Excel file...")
    workbook = load_workbook(output_path)
    
    # Format data sheets
    for sheet_name in tqdm(all_data.keys(), desc="Formatting data sheets"):
        format_data_sheet(workbook, sheet_name)
    
    # Format summary sheet separately with different styling
    if 'Summary' in workbook.sheetnames:
        format_summary_sheet(workbook)
    
    workbook.save(output_path)
    
    print(f"\n✓ Successfully created Excel file: {output_path}")
    print(f"  Total sheets: {len(all_data) + 1} (including Summary)")
    print(f"  Total records: {sum(len(data) for data in all_data.values())}")


def main():
    """Main entry point."""
    # Get project root directory (parent of scripts directory)
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    
    # Define paths
    data_dir = project_root / 'data'
    output_path = project_root / 'artifacts' / 'data_organized.xlsx'
    
    # Create output directory if it doesn't exist
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Check if data directory exists
    if not data_dir.exists():
        print(f"Error: Data directory not found: {data_dir}")
        return
    
    # Organize data
    organize_data_to_xlsx(data_dir, output_path)


if __name__ == '__main__':
    main()

